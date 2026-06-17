from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"


def dataframe_to_excel(df: pd.DataFrame, sheet_name: str = "HaloFest Submissions") -> BytesIO:
    """Create a styled Excel workbook from a DataFrame."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = workbook[sheet_name]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        header_font = Font(color=HEADER_FONT, bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)
    return output
